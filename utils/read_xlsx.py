import openpyxl, sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

wb = openpyxl.load_workbook(r'C:\Users\AbdNahid\Desktop\resume\utils\designation.xlsx')
for sheet in wb.sheetnames:
    ws = wb[sheet]
    print(f'\n=== Sheet: {sheet} ===')
    for row in ws.iter_rows(values_only=True):
        if any(cell is not None for cell in row):
            print('\t'.join(str(c) if c is not None else '' for c in row))
